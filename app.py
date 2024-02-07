from flask import Flask, render_template, request, redirect, url_for, send_file, send_from_directory
import os
import openpyxl
from openpyxl.styles import Font

app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
DOWNLOAD_FOLDER = 'downloads'
ALLOWED_EXTENSIONS = {'xlsx' or 'xls'} 

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['DOWNLOAD_FOLDER'] = DOWNLOAD_FOLDER


def allowed_file(filename):
    # Verifica se a extensão do arquivo é permitida
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def apagar_linhas(lista, min_row=1, max_row=10):
    # Oculta as linhas da planilha dentro do intervalo especificado
    for row_num in range(min_row, max_row):
        lista.row_dimensions[row_num].hidden = True

def desmesclar_celulas(planilha):
    
    for sheet in planilha:
        
        merged_cells_copy = sheet.merged_cells.ranges.copy()
        for merged_cell_range in merged_cells_copy:
            sheet.unmerge_cells(merged_cell_range.coord)

def obter_valor_celula(linha, coluna):
    for cell in linha:
        if coluna == cell.column:  
            if isinstance(cell, openpyxl.cell.MergedCell):
                return cell.value
            return cell.value

def fazer_pedido(arquivo, pagina_produtos):
    # Carrega o workbook 
    livro_excel = openpyxl.load_workbook(arquivo)

    # Desmescla as células
    desmesclar_celulas(livro_excel)

    # Obtém a planilha com os produtos
    pagina_existente = livro_excel[pagina_produtos]

    # Apaga as linhas de 1 a 10 da planilha
    apagar_linhas(pagina_existente, 1, 10)

    # Cria uma nova planilha para os itens necessários
    planilha_necessarios = openpyxl.Workbook()
    nova_pagina = planilha_necessarios.active
    nova_pagina.title = "Itens_Necessarios"

    # Adiciona os cabeçalhos à nova planilha
    nova_pagina.append([cell.value for cell in pagina_existente[10]])

    # Especifica a coluna para a diferença
    coluna_diferenca = 'P'

    # Filtra apenas os itens com estoque menor que a segurança
    itens_necessarios = [row for row in pagina_existente.iter_rows(min_row=11) if row[13].value is not None and row[5].value is not None and row[13].value < row[5].value]

    # Adiciona os itens necessários à nova planilha
    for item in itens_necessarios:
        diferenca = max(0, int(item[5].value) - int(item[13].value))
        nova_pagina.append([obter_valor_celula(item, coluna) for coluna in range(1, len(item) + 1)])
        nova_pagina[f'{coluna_diferenca}{nova_pagina.max_row}'] = diferenca

    # Salva o novo arquivo
    novo_nome_planilha = "Itens_Necessarios.xlsx"
    planilha_necessarios.save(novo_nome_planilha)

    # Move o novo arquivo para a pasta de downloads
    pasta_download = app.config['DOWNLOAD_FOLDER']
    caminho_destino = os.path.join(pasta_download, "Itens_Necessarios.xlsx")
    os.replace(novo_nome_planilha, caminho_destino)

    return caminho_destino

@app.route('/')
def index():
    # Renderiza o template para a página inicial
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return "Nenhum arquivo enviado"
    
    file = request.files['file']

    if file.filename == '':
        return "Nome do arquivo vazio"

    if file and allowed_file(file.filename):
        # Salva o arquivo no servidor
        arquivo = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
        file.save(arquivo)

        # Gera os pedidos e redireciona para a página de sucesso
        novo_arquivo = fazer_pedido(arquivo, "Página1")
        return render_template('success.html')

    return "Formato de arquivo não permitido. Por favor, envie um arquivo xlsx."

@app.route('/success')
def success():
    # Renderiza o template para a página de sucesso
    return render_template('success.html')

@app.route('/downloads/<filename>')
def download(filename):
    file_path = os.path.join(app.config['DOWNLOAD_FOLDER'], filename)
    return send_from_directory(app.config['DOWNLOAD_FOLDER'], filename, as_attachment=True)

if __name__ == '__main__':
    # Inicia o servidor Flask em modo de depuração
    app.run(debug=True)
